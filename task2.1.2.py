import csv
import os
import re

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter


class Vacancy:
    cur_in_rubles = {
        "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
        "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055,
    }

    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.cur_in_rubles[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])


class DataSet:
    def __init__(self, file_name, vacancy_name):
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    @staticmethod
    def increment(dictionary, key, amount):
        if key in dictionary:
            dictionary[key] += amount
        else:
            dictionary[key] = amount

    @staticmethod
    def average(dictionary):
        new_dictionary = {}
        for key, values in dictionary.items():
            new_dictionary[key] = int(sum(values) / len(values))
        return new_dictionary

    def csv(self):
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            header_length = len(header)
            for row in reader:
                if '' not in row and len(row) == header_length:
                    yield dict(zip(header, row))

    def get_statistic(self):
        salary = {}
        SalaryVacNaming = {}
        salary_city = {}
        CountVac = 0

        for dictVac in self.csv():
            vacancy = Vacancy(dictVac)
            self.increment(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.increment(SalaryVacNaming, vacancy.year, [vacancy.salary_average])
            self.increment(salary_city, vacancy.area_name, [vacancy.salary_average])
            CountVac += 1

        NumVac = dict([(key, len(value)) for key, value in salary.items()])
        NumVacancies = dict([(key, len(value)) for key, value in SalaryVacNaming.items()])

        if not SalaryVacNaming:
            SalaryVacNaming = dict([(key, [0]) for key, value in salary.items()])
            NumVacancies = dict([(key, 0) for key, value in NumVac.items()])

        statistic = self.average(salary)
        statistic2 = self.average(SalaryVacNaming)
        statistic3 = self.average(salary_city)

        statistic4 = {}
        for year, salaries in salary_city.items():
            statistic4[year] = round(len(salaries) / CountVac, 4)
        statistic4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in statistic4.items()]))
        statistic4.sort(key=lambda a: a[-1], reverse=True)
        statistic5 = statistic4.copy()
        statistic4 = dict(statistic4)
        statistic3 = list(filter(lambda a: a[0] in list(statistic4.keys()), [(key, value) for key, value in statistic3.items()]))
        statistic3.sort(key=lambda a: a[-1], reverse=True)
        statistic3 = dict(statistic3[:10])
        statistic5 = dict(statistic5[:10])

        return statistic, NumVac, statistic2, NumVacancies, statistic3, statistic5

    @staticmethod
    def print_statistic(stats1, stats2, stats3, stats4, stats5, stats6):
        print('Динамика уровня зарплат по годам: {0}'.format(stats1))
        print('Динамика количества вакансий по годам: {0}'.format(stats2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(stats3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(stats4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(stats5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(stats6))


def cleaner(string):
    return ' '.join(string.split())


def reading(file_name: str):
    if os.stat(file_name).st_size == 0:
        print("Пустой файл")
        return -1, -1

    with open(file_name, 'r', encoding='utf-8-sig') as file_:
        lines = list(csv.reader(file_))
        titles = lines[0]
        rows = [line for line in lines[1:] if '' not in line and len(line) == len(titles)]
        return titles, rows


def clean_t(value):
    result = re.sub(re.compile('<.*?>'), '', value)
    if '\n' in result:
        return [cleaner(i) for i in result.split('\n')]
    return cleaner(result)


def filing(name_of_list, read):
    result = []
    for line in read:
        saver = []
        for elem in line:
            saver.append(clean_t(elem))
        result.append(dict(zip(name_of_list, saver)))
    return result


class InputConnect:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')


        dataset = DataSet(self.file_name, self.vacancy_name)
        statistic1, statistic2, statistic3, statistic4, statistic5, statistic6 = dataset.get_statistic()
        dataset.print_statistic(statistic1, statistic2, statistic3, statistic4, statistic5, statistic6)

        report = Report(self.vacancy_name, statistic1, statistic2, statistic3, statistic4, statistic5, statistic6)
        report.generate_excel()
        report.save('report.xlsx')
        report.CreatePicture()


class Report:
    def __init__(self, NameVac, statistic1, statistic2, statistic3, statistic4, statistic5, statistic6):
        self.wb = Workbook()
        self.NameVac = NameVac
        self.stats1 = statistic1
        self.stats2 = statistic2
        self.stats3 = statistic3
        self.stats4 = statistic4
        self.stats5 = statistic5
        self.stats6 = statistic6

    def generate_excel(self):
        ws1 = self.wb.active
        ws1.title = 'Статистика по годам'
        ws1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.NameVac, 'Количество вакансий',
                    'Количество вакансий - ' + self.NameVac])
        for year in self.stats1.keys():
            ws1.append([year, self.stats1[year], self.stats3[year], self.stats2[year], self.stats4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.NameVac, ' Количество вакансий',
                 ' Количество вакансий - ' + self.NameVac]]
        column_widths = []
        for column in data:
            for i, cell in enumerate(column):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, x in enumerate(column_widths, 1):  # ,1 to start at 1
            ws1.column_dimensions[get_column_letter(i)].width = x + 2

        data = []
        data.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city1, value1), (city2, value2) in zip(self.stats5.items(), self.stats6.items()):
            data.append([city1, value1, '', city2, value2])
        ws2 = self.wb.create_sheet('Статистика по городам')
        for column in data:
            ws2.append(column)

        column_widths = []
        for column in data:
            for i, cell in enumerate(column):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, x in enumerate(column_widths, 1):  # ,1 to start at 1
            ws2.column_dimensions[get_column_letter(i)].width = x + 2

        border = Font(bold=True)
        for col in 'ABCDE':
            ws1[col + '1'].font = border
            ws2[col + '1'].font = border

        for index, _ in enumerate(self.stats5):
            ws2['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        for column in range(len(data)):
            for col in 'ABDE':
                ws2[col + str(column + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        for column, _ in enumerate(self.stats1):
            for col in 'ABCDE':
                ws1[col + str(column + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

    def CreatePicture(self):
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)

        bar1 = ax1.bar(np.array(list(self.stats1.keys())) - 0.4, self.stats1.values(), width=0.4)
        bar2 = ax1.bar(np.array(list(self.stats1.keys())), self.stats3.values(), width=0.4)
        ax1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        ax1.grid(axis='y')
        ax1.legend((bar1[0], bar2[0]), ('средняя з/п', 'з/п ' + self.NameVac.lower()), prop={'size': 8})
        ax1.set_xticks(np.array(list(self.stats1.keys())) - 0.2, list(self.stats1.keys()), rotation=90)
        ax1.xaxis.set_tick_params(labelsize=8)
        ax1.yaxis.set_tick_params(labelsize=8)

        ax2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        bar1 = ax2.bar(np.array(list(self.stats2.keys())) - 0.4, self.stats2.values(), width=0.4)
        bar2 = ax2.bar(np.array(list(self.stats2.keys())), self.stats4.values(), width=0.4)
        ax2.legend((bar1[0], bar2[0]), ('Количество вакансий', 'Количество вакансий\n' + self.NameVac.lower()),
                   prop={'size': 8})
        ax2.set_xticks(np.array(list(self.stats2.keys())) - 0.2, list(self.stats2.keys()), rotation=90)
        ax2.grid(axis='y')
        ax2.xaxis.set_tick_params(labelsize=8)
        ax2.yaxis.set_tick_params(labelsize=8)

        ax3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        ax3.barh(list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.stats5.keys()))]),
                 list(reversed(list(self.stats5.values()))), color='blue', height=0.5, align='center')
        ax3.yaxis.set_tick_params(labelsize=6)
        ax3.xaxis.set_tick_params(labelsize=8)
        ax3.grid(axis='x')

        ax4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum([value for value in self.stats6.values()])
        ax4.pie(list(self.stats6.values()) + [other], labels=list(self.stats6.keys()) + ['Другие'],
                textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph.png')

    def save(self, filename):
        self.wb.save(filename=filename)



InputConnect()
