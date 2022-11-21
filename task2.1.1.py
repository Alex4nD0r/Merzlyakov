import csv
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter


class Vacancy:
    currency_to_rub = {
        "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
        "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055,
    }

    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])


class DataSet:
    def __init__(self, file_name, vacancy_name):
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    @staticmethod
    def increment(dictionary, key, amount):
        if key in dictionary:
            dictionary[key] += amount
        else:
            dictionary[key] = amount

    @staticmethod
    def average(dictionary):
        new_dictionary = {}
        for key, values in dictionary.items():
            new_dictionary[key] = int(sum(values) / len(values))
        return new_dictionary

    def csv(self):
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            header_length = len(header)
            for row in reader:
                if '' not in row and len(row) == header_length:
                    yield dict(zip(header, row))

    def get_statistic(self):
        salary = {}
        name_salary_vac = {}
        CitySalary = {}
        count_of_vacancies = 0

        for i in self.csv():
            vacancy = Vacancy(i)
            self.increment(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.increment(name_salary_vac, vacancy.year, [vacancy.salary_average])
            self.increment(CitySalary, vacancy.area_name, [vacancy.salary_average])
            count_of_vacancies += 1

        vacancies_number = dict([(key, len(value)) for key, value in salary.items()])
        numberVacancy = dict([(key, len(value)) for key, value in name_salary_vac.items()])

        if not name_salary_vac:
            name_salary_vac = dict([(key, [0]) for key, value in salary.items()])
            numberVacancy = dict([(key, 0) for key, value in vacancies_number.items()])

        statistic = self.average(salary)
        statistic2 = self.average(name_salary_vac)
        statistic3 = self.average(CitySalary)

        statistic4 = {}
        for year, salaries in CitySalary.items():
            statistic4[year] = round(len(salaries) / count_of_vacancies, 4)
        statistic4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in statistic4.items()]))
        statistic4.sort(key=lambda a: a[-1], reverse=True)
        statistic5 = statistic4.copy()
        statistic4 = dict(statistic4)
        statistic3 = list(filter(lambda a: a[0] in list(statistic4.keys()), [(key, value) for key, value in statistic3.items()]))
        statistic3.sort(key=lambda a: a[-1], reverse=True)
        statistic3 = dict(statistic3[:10])
        statistic5 = dict(statistic5[:10])

        return statistic, vacancies_number, statistic2, numberVacancy, statistic3, statistic5

    @staticmethod
    def Stat_print(statistic1, statistic2, statistic3, statistic4, statistic5, statistic6):
        print('Динамика уровня зарплат по годам: {0}'.format(statistic1))
        print('Динамика количества вакансий по годам: {0}'.format(statistic2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(statistic3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(statistic4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(statistic5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(statistic6))


def methodForInputConnect():
    global InputConnect

    class InputConnect:
        def __init__(self):
            self.file_name = input('Введите название файла: ')
            self.vacancy_name = input('Введите название профессии: ')

            dataset = DataSet(self.file_name, self.vacancy_name)
            stats1, stats2, stats3, stats4, stats5, stats6 = dataset.get_statistic()
            dataset.Stat_print(stats1, stats2, stats3, stats4, stats5, stats6)

            report = Report(self.vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6)
            report.generate_excel()


methodForInputConnect()


def cleaner(string):
    return ' '.join(string.split())


def reading(file_name: str):
    if os.stat(file_name).st_size == 0:
        print("Пустой файл")
        return -1, -1

    with open(file_name, 'r', encoding='utf-8-sig') as file_:
        lines = list(csv.reader(file_))
        titles = lines[0]
        rows = [line for line in lines[1:] if '' not in line and len(line) == len(titles)]
        return titles, rows


def clean_t(value):
    result = re.sub(re.compile('<.*?>'), '', value)
    if '\n' in result:
        return [cleaner(i) for i in result.split('\n')]
    return cleaner(result)


def filing(name_of_list, read):
    result = []
    for line in read:
        saver = []
        for elem in line:
            saver.append(clean_t(elem))
        result.append(dict(zip(name_of_list, saver)))
    return result


class Report:
    def __init__(self, vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6):
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.stats1 = stats1
        self.stats2 = stats2
        self.stats3 = stats3
        self.stats4 = stats4
        self.stats5 = stats5
        self.stats6 = stats6

    def generate_excel(self):
        ws1 = self.wb.active
        ws1.title = 'Статистика по годам'
        ws1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий',
                    'Количество вакансий - ' + self.vacancy_name])
        for year in self.stats1.keys():
            ws1.append([year, self.stats1[year], self.stats3[year], self.stats2[year], self.stats4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancy_name]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            ws1.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = []
        data.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city1, value1), (city2, value2) in zip(self.stats5.items(), self.stats6.items()):
            data.append([city1, value1, '', city2, value2])
        ws2 = self.wb.create_sheet('Статистика по городам')
        for row in data:
            ws2.append(row)

        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            ws2.column_dimensions[get_column_letter(i)].width = column_width + 2

        font_bold = Font(bold=True)
        for col in 'ABCDE':
            ws1[col + '1'].font = font_bold
            ws2[col + '1'].font = font_bold

        for index, _ in enumerate(self.stats5):
            ws2['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for col in 'ABDE':
                ws2[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        self.stats1[1] = 1
        for row, _ in enumerate(self.stats1):
            for col in 'ABCDE':
                ws1[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        self.wb.save('report.xlsx')


InputConnect()
